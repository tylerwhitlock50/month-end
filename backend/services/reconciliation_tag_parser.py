"""
Reconciliation Tag Parser Service

This service extracts reconciliation values from Excel/CSV files by:
1. Scanning all cells for tags matching pattern TB-{period_id}-{account_number}
2. Extracting the numeric value from the cell immediately to the left
3. Supporting multiple tags in a single file for bulk reconciliation

Supported file formats:
- .xlsx (Excel 2007+)
- .xls (Excel 97-2003)
- .csv (Comma-separated values)
"""

import csv
import io
import re
from decimal import Decimal, InvalidOperation
from typing import Dict, List, Optional, Tuple
import openpyxl
import xlrd


class ReconciliationTagParser:
    """Parser for extracting reconciliation values from spreadsheet files."""
    
    TAG_PATTERN = re.compile(r'TB-(\d+)-([A-Za-z0-9\-\.]+)')
    
    def __init__(self, file_bytes: bytes, filename: str):
        """
        Initialize parser with file data.
        
        Args:
            file_bytes: Raw file content
            filename: Original filename (used to determine file type)
        """
        self.file_bytes = file_bytes
        self.filename = filename.lower()
        self.tags_found: Dict[str, Decimal] = {}
        self.errors: List[str] = []
    
    def parse(self) -> Dict[str, Decimal]:
        """
        Parse the file and extract all reconciliation tag values.
        
        Returns:
            Dictionary mapping tag strings to extracted Decimal values
            Example: {"TB-1-1000": Decimal("5000.00"), "TB-1-2000": Decimal("3500.00")}
        """
        if self.filename.endswith('.xlsx'):
            return self._parse_xlsx()
        elif self.filename.endswith('.xls'):
            return self._parse_xls()
        elif self.filename.endswith('.csv'):
            return self._parse_csv()
        else:
            self.errors.append(f"Unsupported file format: {self.filename}")
            return {}
    
    def _parse_xlsx(self) -> Dict[str, Decimal]:
        """Parse Excel 2007+ (.xlsx) files using openpyxl."""
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(self.file_bytes), data_only=True)
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                self._scan_excel_sheet(sheet)
            
            return self.tags_found
        
        except Exception as e:
            self.errors.append(f"Error parsing XLSX file: {str(e)}")
            return {}
    
    def _parse_xls(self) -> Dict[str, Decimal]:
        """Parse Excel 97-2003 (.xls) files using xlrd."""
        try:
            workbook = xlrd.open_workbook(file_contents=self.file_bytes)
            
            for sheet_idx in range(workbook.nsheets):
                sheet = workbook.sheet_by_index(sheet_idx)
                self._scan_xlrd_sheet(sheet)
            
            return self.tags_found
        
        except Exception as e:
            self.errors.append(f"Error parsing XLS file: {str(e)}")
            return {}
    
    def _parse_csv(self) -> Dict[str, Decimal]:
        """
        Parse CSV files.
        
        For CSV files, we scan each cell for tags and look for the value
        in the previous column of the same row.
        """
        try:
            text = self.file_bytes.decode('utf-8')
            reader = csv.reader(io.StringIO(text))
            
            for row_data in reader:
                self._scan_csv_row(row_data)
            
            return self.tags_found
        
        except UnicodeDecodeError:
            # Try with latin-1 encoding if UTF-8 fails
            try:
                text = self.file_bytes.decode('latin-1')
                reader = csv.reader(io.StringIO(text))
                
                for row_data in reader:
                    self._scan_csv_row(row_data)
                
                return self.tags_found
            
            except Exception as e:
                self.errors.append(f"Error parsing CSV file: {str(e)}")
                return {}
        
        except Exception as e:
            self.errors.append(f"Error parsing CSV file: {str(e)}")
            return {}
    
    def _scan_excel_sheet(self, sheet):
        """Scan an openpyxl worksheet for tags and extract values."""
        for row in sheet.iter_rows():
            for col_idx, cell in enumerate(row):
                if cell.value is None:
                    continue
                
                cell_value = str(cell.value).strip()
                match = self.TAG_PATTERN.search(cell_value)
                
                if match:
                    tag = match.group(0)
                    
                    # Get value from cell to the left (same row, previous column)
                    if col_idx > 0:
                        left_cell = row[col_idx - 1]
                        extracted_value = self._extract_numeric_value(left_cell.value)
                        
                        if extracted_value is not None:
                            if tag in self.tags_found:
                                self.errors.append(f"Duplicate tag found: {tag}")
                            else:
                                self.tags_found[tag] = extracted_value
                        else:
                            self.errors.append(
                                f"Could not extract numeric value for tag {tag} "
                                f"from cell value: {left_cell.value}"
                            )
                    else:
                        self.errors.append(
                            f"Tag {tag} found in first column with no value to the left"
                        )
    
    def _scan_xlrd_sheet(self, sheet):
        """Scan an xlrd worksheet for tags and extract values."""
        for row_idx in range(sheet.nrows):
            for col_idx in range(sheet.ncols):
                cell = sheet.cell(row_idx, col_idx)
                
                if cell.ctype == xlrd.XL_CELL_EMPTY:
                    continue
                
                cell_value = str(cell.value).strip()
                match = self.TAG_PATTERN.search(cell_value)
                
                if match:
                    tag = match.group(0)
                    
                    # Get value from cell to the left (same row, previous column)
                    if col_idx > 0:
                        left_cell = sheet.cell(row_idx, col_idx - 1)
                        extracted_value = self._extract_numeric_value(left_cell.value)
                        
                        if extracted_value is not None:
                            if tag in self.tags_found:
                                self.errors.append(f"Duplicate tag found: {tag}")
                            else:
                                self.tags_found[tag] = extracted_value
                        else:
                            self.errors.append(
                                f"Could not extract numeric value for tag {tag} "
                                f"from cell value: {left_cell.value}"
                            )
                    else:
                        self.errors.append(
                            f"Tag {tag} found in first column with no value to the left"
                        )
    
    def _scan_csv_row(self, row_data: List[str]):
        """Scan a CSV row for tags and extract values from the previous column."""
        for col_idx, cell_value in enumerate(row_data):
            if not cell_value:
                continue
            
            cell_value = cell_value.strip()
            match = self.TAG_PATTERN.search(cell_value)
            
            if match:
                tag = match.group(0)
                
                # Get value from previous column in the same row
                if col_idx > 0:
                    left_value = row_data[col_idx - 1]
                    extracted_value = self._extract_numeric_value(left_value)
                    
                    if extracted_value is not None:
                        if tag in self.tags_found:
                            self.errors.append(f"Duplicate tag found: {tag}")
                        else:
                            self.tags_found[tag] = extracted_value
                    else:
                        self.errors.append(
                            f"Could not extract numeric value for tag {tag} "
                            f"from cell value: {left_value}"
                        )
                else:
                    self.errors.append(
                        f"Tag {tag} found in first column with no value to the left"
                    )
    
    def _extract_numeric_value(self, value) -> Optional[Decimal]:
        """
        Extract a numeric value from a cell, handling various formats.
        
        Args:
            value: Cell value (could be number, string, formula result, etc.)
        
        Returns:
            Decimal value if extraction successful, None otherwise
        """
        if value is None:
            return None
        
        # Handle numeric types directly
        if isinstance(value, (int, float)):
            try:
                return Decimal(str(value))
            except (InvalidOperation, ValueError):
                return None
        
        # Handle string values
        if isinstance(value, str):
            # Remove common formatting characters
            cleaned = value.strip()
            
            # Remove currency symbols, commas, and spaces
            cleaned = re.sub(r'[$,\s€£¥]', '', cleaned)
            
            # Handle parentheses for negative numbers (accounting format)
            if cleaned.startswith('(') and cleaned.endswith(')'):
                cleaned = '-' + cleaned[1:-1]
            
            try:
                return Decimal(cleaned)
            except (InvalidOperation, ValueError):
                return None
        
        return None
    
    def get_errors(self) -> List[str]:
        """Return list of errors encountered during parsing."""
        return self.errors


def extract_reconciliation_values(
    file_bytes: bytes,
    filename: str,
    period_id: Optional[int] = None
) -> Tuple[Dict[str, Decimal], List[str]]:
    """
    Convenience function to extract reconciliation values from a file.
    
    Args:
        file_bytes: Raw file content
        filename: Original filename
        period_id: Optional period ID to filter tags (only return tags for this period)
    
    Returns:
        Tuple of (tags_dict, errors_list)
        - tags_dict: Dictionary mapping tag strings to Decimal values
        - errors_list: List of error messages encountered during parsing
    """
    parser = ReconciliationTagParser(file_bytes, filename)
    tags = parser.parse()
    
    # Filter by period_id if provided
    if period_id is not None:
        filtered_tags = {
            tag: value
            for tag, value in tags.items()
            if tag.startswith(f'TB-{period_id}-')
        }
        return filtered_tags, parser.get_errors()
    
    return tags, parser.get_errors()

